from django.apps.registry import apps
from django.shortcuts import render
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
# settings de la app
from django.conf import settings
# reverse url
from django.urls import reverse
from django.http import HttpResponseRedirect

# propios
from configuraciones.models import Cajas, Puntos, Sucursales, Ciudades, Almacenes, Lineas
from status.models import Status

# para los usuarios
from utils.permissions import get_user_permission_operation, get_permissions_user
from utils.dates_functions import get_date_system, get_date_show, get_date_to_db

# clases por modulo
from controllers.reportes.ReportesController import ReportesController
from controllers.ventas.VentasController import VentasController

import os
# xls
import openpyxl

# reportes
import io
from django.http import FileResponse, HttpResponse
from reportlab.pdfgen import canvas

from reportes.rptArqueoCaja import rptArqueoCaja
from reportes.rptIngresosCaja import rptIngresosCaja
from reportes.rptEgresosCaja import rptEgresosCaja
from reportes.rptMovimientosCaja import rptMovimientosCaja
from reportes.rptIngresoProductos import rptIngresoProductos
from reportes.rptSalidaProductos import rptSalidaProductos
from reportes.rptMovimientoProductos import rptMovimientoProductos
from reportes.rptVentas import rptVentas
from reportes.rptStockProductos import rptStockProductos

# reportes controller
reportes_controller = ReportesController()


@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def reportes_index(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    status_activo = reportes_controller.status_activo

    # operaciones
    if 'operation_x' in request.POST.keys():
        # verificamos operacion valida
        operation = request.POST['operation_x']
        if not operation in ['', 'arqueo_caja', 'ingresos_caja', 'egresos_caja', 'movimientos_caja',
                             'buscar_sucursal', 'buscar_solo_sucursal', 'buscar_sucursal_punto', 'buscar_caja', 'buscar_punto',
                             'buscar_sucursal_almacen', 'buscar_almacen',
                             'ingreso_productos', 'salida_productos', 'movimiento_productos',
                             'ventas', 'stock_productos']:

            return render(request, 'pages/without_permission.html')

        # stock productos
        if operation == 'stock_productos':
            respuesta = stock_productos(request)
            if not type(respuesta) == bool:
                return respuesta

        # ventas
        if operation == 'ventas':
            respuesta = ventas(request)
            if not type(respuesta) == bool:
                return respuesta

        # inventarios
        if operation == 'ingreso_productos':
            respuesta = ingreso_productos(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'salida_productos':
            respuesta = salida_productos(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'movimiento_productos':
            respuesta = movimiento_productos(request)
            if not type(respuesta) == bool:
                return respuesta

        # cajas
        if operation == 'movimientos_caja':
            respuesta = movimientos_caja(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'egresos_caja':
            respuesta = egresos_caja(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'ingresos_caja':
            respuesta = ingresos_caja(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'arqueo_caja':
            respuesta = arqueo_caja(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'buscar_sucursal':
            ciudad_id = request.POST['ciudad'].strip()
            ciudad = Ciudades.objects.get(pk=ciudad_id)
            lista_sucursales = Sucursales.objects.filter(status_id=status_activo, ciudad_id=ciudad).order_by('sucursal')

            context_s = {
                'lista_sucursales': lista_sucursales,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_sucursal.html', context_s)

        if operation == 'buscar_solo_sucursal':
            ciudad_id = request.POST['ciudad'].strip()
            ciudad = Ciudades.objects.get(pk=ciudad_id)
            lista_sucursales = Sucursales.objects.filter(status_id=status_activo, ciudad_id=ciudad).order_by('sucursal')

            context_s = {
                'lista_sucursales': lista_sucursales,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_solo_sucursal.html', context_s)

        if operation == 'buscar_sucursal_punto':
            ciudad_id = request.POST['ciudad'].strip()
            ciudad = Ciudades.objects.get(pk=ciudad_id)
            lista_sucursales = Sucursales.objects.filter(status_id=status_activo, ciudad_id=ciudad).order_by('sucursal')

            context_s = {
                'lista_sucursales': lista_sucursales,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_sucursal_punto.html', context_s)

        if operation == 'buscar_sucursal_almacen':
            ciudad_id = request.POST['ciudad'].strip()
            ciudad = Ciudades.objects.get(pk=ciudad_id)
            lista_sucursales = Sucursales.objects.filter(status_id=status_activo, ciudad_id=ciudad).order_by('sucursal')

            context_s = {
                'lista_sucursales': lista_sucursales,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_sucursal_almacen.html', context_s)

        if operation == 'buscar_caja':
            sucursal_id = request.POST['sucursal'].strip()
            sucursal = Sucursales.objects.get(pk=sucursal_id)
            filtro = {}
            filtro['status_id'] = status_activo
            filtro['punto_id__sucursal_id'] = sucursal

            lista_cajas = Cajas.objects.select_related('punto_id').filter(**filtro).order_by('punto_id__punto', 'caja')

            context_s = {
                'lista_cajas': lista_cajas,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_caja.html', context_s)

        if operation == 'buscar_punto':
            sucursal_id = request.POST['sucursal'].strip()
            sucursal = Sucursales.objects.get(pk=sucursal_id)
            filtro = {}
            filtro['status_id'] = status_activo
            filtro['sucursal_id'] = sucursal

            lista_puntos = Puntos.objects.select_related('sucursal_id').filter(**filtro).order_by('punto')

            context_s = {
                'lista_puntos': lista_puntos,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_punto.html', context_s)

        if operation == 'buscar_almacen':
            sucursal_id = request.POST['sucursal'].strip()
            sucursal = Sucursales.objects.get(pk=sucursal_id)
            filtro = {}
            filtro['status_id'] = status_activo
            filtro['sucursal_id'] = sucursal

            lista_almacenes = Almacenes.objects.select_related('sucursal_id').filter(**filtro).order_by('almacen')

            context_s = {
                'lista_almacenes': lista_almacenes,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_almacen.html', context_s)

    # verificamos mensajes
    if 'nuevo_mensaje' in request.session.keys():
        messages.add_message(request, messages.SUCCESS, request.session['nuevo_mensaje'])
        del request.session['nuevo_mensaje']
        request.session.modified = True

    # datos por defecto
    context = {
        'permisos': permisos,
        'url_main': '',
        # 'js_file': reportes_controller.modulo_session,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': '',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/reportes.html', context)


# arqueo de caja
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def arqueo_caja(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        # reporte de arqueo de caja
        fecha = get_date_to_db(fecha=request.POST['fecha'], formato_ori='dd-MMM-yyyy', formato='yyyy-mm-dd')

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptArqueoCaja(buffer, int(request.POST['id']), fecha)

            buffer.seek(0)
            return FileResponse(buffer, filename='arqueo_caja.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    # lista de cajas de la sucursal
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'arqueo_caja',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/arqueo_caja.html', context)


# ingresos a caja
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def ingresos_caja(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptIngresosCaja(buffer, request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            return FileResponse(buffer, filename='ingresos_caja.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Ciudad', 'Sucursal', 'Punto', 'Caja', 'Operacion', 'Fecha', 'concepto', 'monto', 'codigo', 'estado'))
            datos_reporte = reportes_controller.datos_ingreso_caja(request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['punto'], dato['caja'], dato['operacion'], dato['fecha'], dato['concepto'], dato['monto'], dato['tipo_moneda'], dato['estado_txt']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=ingresos_caja.xlsx"
            wb.save(response)
            return response
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    # lista ciudades
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    # lista de cajas de la sucursal
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'fecha_actual': fecha_actual,
        'lista_ciudades': lista_ciudades,
        'permisos': permisos,
        'url_actual': '',

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'ingresos_caja',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/ingresos_caja.html', context)


# egresos caja
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def egresos_caja(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptEgresosCaja(buffer, request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='egresos_caja.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Ciudad', 'Sucursal', 'Punto', 'Caja', 'Operacion', 'Fecha', 'concepto', 'monto', 'codigo', 'estado'))
            datos_reporte = reportes_controller.datos_egreso_caja(request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['punto'], dato['caja'], dato['operacion'], dato['fecha'], dato['concepto'], dato['monto'], dato['tipo_moneda'], dato['estado_txt']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=egresos_caja.xlsx"
            wb.save(response)
            return response
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    # lista ciudades
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    # lista de cajas
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'egresos_caja',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/egresos_caja.html', context)


# movimientos caja
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def movimientos_caja(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptMovimientosCaja(buffer, request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='movimientos_caja.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Ciudad', 'Sucursal', 'Fecha', 'Concepto', 'Monto', 'Codigo', 'Caja1', 'Punto1', 'Caja2', 'Punto2', 'estado'))
            datos_reporte = reportes_controller.datos_movimientos_caja(request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['fecha'], dato['concepto'], dato['monto'], dato['tipo_moneda'], dato['caja1'], dato['punto1'], dato['caja2'], dato['punto2'], dato['estado_txt']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=movimientos_caja.xlsx"
            wb.save(response)
            return response
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'movimientos_caja',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/movimientos_caja.html', context)


# ingreso de productos
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def ingreso_productos(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptIngresoProductos(buffer, request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='ingreso_productos.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "IngresoProductos"
            hoja.append(('Ciudad', 'Sucursal', 'Almacen', 'Linea', 'Producto', 'Cantidad'))
            datos_reporte = reportes_controller.datos_ingreso_productos(request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['almacen'], dato['linea'], dato['producto'], dato['cantidad']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=ingreso_productos.xlsx"
            wb.save(response)
            return response

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    lista_almacenes = reportes_controller.lista_almacenes_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_almacenes': lista_almacenes,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'ingreso_productos',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/ingreso_productos.html', context)


# salida de productos
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def salida_productos(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptSalidaProductos(buffer, request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='salida_productos.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "SalidaProductos"
            hoja.append(('Ciudad', 'Sucursal', 'Almacen', 'Linea', 'Producto', 'Cantidad'))
            datos_reporte = reportes_controller.datos_salida_productos(request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['almacen'], dato['linea'], dato['producto'], dato['cantidad']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=salida_productos.xlsx"
            wb.save(response)
            return response

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    lista_almacenes = reportes_controller.lista_almacenes_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_almacenes': lista_almacenes,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'salida_productos',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/salida_productos.html', context)


# movimiento de productos
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def movimiento_productos(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptMovimientoProductos(buffer, request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='movimiento_productos.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "MovimientoProductos"
            hoja.append(('Ciudad', 'Sucursal', 'Almacenes', 'Linea', 'Producto', 'Cantidad'))
            datos_reporte = reportes_controller.datos_movimiento_productos(request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['almacen'], dato['linea'], dato['producto'], dato['cantidad']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=movimiento_productos.xlsx"
            wb.save(response)
            return response

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    lista_almacenes = reportes_controller.lista_almacenes_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_almacenes': lista_almacenes,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'movimiento_productos',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/movimiento_productos.html', context)


# ventas
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def ventas(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        punto = int(request.POST['punto'].strip())
        anulados = request.POST['anulados'].strip()
        preventa = request.POST['preventa'].strip()
        venta = request.POST['venta'].strip()
        salida = request.POST['salida'].strip()
        vuelta = request.POST['vuelta'].strip()
        finalizado = request.POST['finalizado'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptVentas(buffer, request.user, ciudad, sucursal, punto, fecha_ini, fecha_fin, anulados, preventa, venta, salida, vuelta, finalizado)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='ventas.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        punto = int(request.POST['punto'].strip())
        anulados = request.POST['anulados'].strip()
        preventa = request.POST['preventa'].strip()
        venta = request.POST['venta'].strip()
        salida = request.POST['salida'].strip()
        vuelta = request.POST['vuelta'].strip()
        finalizado = request.POST['finalizado'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Ventas"
            hoja.append(('Ciudad', 'Sucursal', 'Punto', 'Cliente', 'Tipo', 'Subtotal', 'Descuento', 'Adicional', 'Total', 'Ingresos', 'Numero', 'Fecha'))
            datos_reporte = reportes_controller.datos_ventas(request.user, ciudad, sucursal, punto, fecha_ini, fecha_fin, anulados, preventa, venta, salida, vuelta, finalizado)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['punto'], dato['cliente'], dato['tipo'], dato['subtotal'], dato['descuento'],
                            dato['adicional'], dato['total'] + dato['adicional'], dato['ingresos_caja'], dato['numero_contrato'], dato['fecha']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=ventas.xlsx"
            wb.save(response)
            return response

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    lista_puntos = reportes_controller.lista_puntos_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_puntos': lista_puntos,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'ventas',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/ventas.html', context)


# stock productos
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def stock_productos(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        linea = request.POST['linea'].strip()
        almacen = request.POST['almacen'].strip()
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptStockProductos(buffer, request.user, linea, almacen, fecha_ini, fecha_fin)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='stock_productos.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        linea = request.POST['linea'].strip()
        almacen = request.POST['almacen'].strip()
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "StockProductos"
            hoja.append(('Almacen', 'Linea', 'Producto', 'Cantidad'))
            #datos_reporte = reportes_controller.datos_stock_productos(request.user, linea, almacen)
            venta_controller = VentasController()
            almacen1 = apps.get_model('configuraciones', 'Almacenes').objects.get(pk=1)
            datos_stock = venta_controller.stock_productos(fecha_entrega=fecha_ini, fecha_devolucion=fecha_fin, formato='dd-MMM-yyyy')
            datos_reporte = []

            lista_productos = apps.get_model('productos', 'Productos').objects.filter(status_id=venta_controller.status_activo).order_by('linea_id__linea', 'producto')
            datos_reporte = []
            valor_linea = int(linea)
            for li_pro in lista_productos:
                entra = 'si'
                if valor_linea != 0:
                    if valor_linea != li_pro.linea_id.linea_id:
                        entra = 'no'
                if entra == 'si':
                    dato_add = {}
                    dato_add['linea'] = li_pro.linea_id.linea
                    dato_add['producto'] = li_pro.producto
                    cantidad = datos_stock[li_pro.producto_id]
                    dato_add['cantidad'] = cantidad

                    datos_reporte.append(dato_add)
                    hoja.append((almacen1.almacen, dato_add['linea'], dato_add['producto'], dato_add['cantidad']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=stock_productos.xlsx"
            wb.save(response)
            return response

    # lista de cajas de la sucursal
    lista_lineas = Lineas.objects.filter(status_id=status_activo).order_by('linea')

    # lista de almacenes
    lista_almacenes = apps.get_model('configuraciones', 'Almacenes').objects.filter(status_id=status_activo).order_by('almacen_id')
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_lineas': lista_lineas,
        'lista_almacenes': lista_almacenes,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'stock_productos',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/stock_productos.html', context)
